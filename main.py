import requests
import json
import optparse
import openpyxl
import sys

BASE_URL = "https://api.meraki.com/api/v1/"


def get_network_objs(network):
    url = BASE_URL+"organizations/{}/networkObjects".format(network['org_id'])
    payload = {}
    headers = {'X-Cisco-Meraki-API-Key': network['api_key']}
    response = requests.request("GET", url, headers=headers, data=payload)
    return response.json()


def get_network_object_group(network):
    url = BASE_URL+"organizations/{}/networkObjectGroups/".format(network['org_id'])
    payload = {}
    headers = {'X-Cisco-Meraki-API-Key': network['api_key']}
    response = requests.request("GET", url, headers=headers, data=payload)
    return response.json()


def post_network_obj(network, nw_obj):
    url = BASE_URL+"organizations/{}/networkObjects".format(network['org_id'])
    payload = {"name": nw_obj["name"], "type": nw_obj["type"], "value": nw_obj["value"], "networkObjectGroupIds":[]}
    headers = {'X-Cisco-Meraki-API-Key': network['api_key']}
    response = requests.request("POST", url, headers=headers, data=payload)
    print("Network Object: {}\n\tResponse:".format(nw_obj['name']), response)
    return response


def post_network_object_group(network, nw_obj_grp):
    url = BASE_URL+"organizations/{}/networkObjectGroups/".format(network['org_id'])
    payload = {"name": nw_obj_grp['name'], "networkObjectIds": nw_obj_grp['networkObjectIds']}
    str_payload = str(payload).replace("'", '"')
    headers = {
        'X-Cisco-Meraki-API-Key': network['api_key'],
        'Content-Type': "application/json"
    }
    response = requests.request("POST", url, headers=headers, data=str_payload)
    print("Object Group: {}\n\tResponse:".format(nw_obj_grp['name']), response)
    return response


def post_fw_rule(network, fw_rule_list):
    url = BASE_URL+"networks/{}/appliance/firewall/l3FirewallRules".format(network['net_id'])
    payload = {"rules": fw_rule_list}
    str_payload = json.dumps(payload)
    headers = {
        'X-Cisco-Meraki-API-Key': network['api_key'],
        'Content-Type': "application/json"
    }
    response = requests.request("PUT", url, headers=headers, data=str_payload)
    print("FW Rule Reponse:", response)
    return response


def cli_args():
    """Reads the CLI options provided and returns them using the OptionParser
    Will return the Values as a dictionary"""
    parser = optparse.OptionParser()
    parser.add_option('-v','--verbose',
                      dest="verbose",
                      default=False,
                      action="store_true",
                      help="Enable Verbose Output"
                      )
    parser.add_option('-i', '--input_file',
                      dest="input_file",
                      default="Meraki - Network Object Firewall Rules.xlsx",
                      action="store",
                      help="Organization ID"
                      )
    parser.add_option('-k','--api_key',
                      dest="api_key",
                      action="store",
                      help="API Key"
                      )
    parser.add_option('-o','--org_id',
                      dest="org_id",
                      default="",
                      action="store",
                      help="Organization ID"
                      )
    parser.add_option('-n', '--net_id',
                      dest="net_id",
                      action="store",
                      help="Network ID"
                      )
    parser.add_option('-f','--output_file',
                      dest="output_file",
                      action="store",
                      help="Output file name of excel sheet"
                      )

    options, remainder = parser.parse_args()
    # Utilizing the vars() method we can return the options as a dictionary
    return vars(options)


def open_xls(xls_input_file_name):
    """
    Returns the WorkBook of specified Name
    """
    try:
        return openpyxl.load_workbook(xls_input_file_name, data_only=True)
    except Exception as e:
        print(e)
        print("Please ensure the file exists or the correct filename was entered when utilizing the \"-i\" argument.")


def rw_cell(sheet_obj, row, column, write=False, value=""):
    """
    Either writes or reads to/from a cell.
    """
    if write:
        sheet_obj.cell(row=row, column=column).value = value
        return None
    return sheet_obj.cell(row=row, column=column).value


def wb_read_network_objects(sheet_obj):
    return_dict = []
    for i in range(2,sheet_obj.max_row+1):
        t_dict = {
            'name': rw_cell(sheet_obj, i, 1).strip(),
            'type': rw_cell(sheet_obj, i, 2),
            'value': rw_cell(sheet_obj, i, 3)
        }
        return_dict.append(t_dict)
    return return_dict


def wb_read_network_object_groups(sheet_obj):
    return_list = []
    for i in range(2,sheet_obj.max_row+1):
        t_dict = {
            'name': rw_cell(sheet_obj, i, 1),
            'networkObjectNames': rw_cell(sheet_obj, i, 2),
            'networkObjectIds': [],
            'groupID': '',
            'errors': []
        }
        return_list.append(t_dict)
    return return_list


def wb_read_fw_rules(sheet_obj):
    return_list = []
    for i in range(2,sheet_obj.max_row+1):
        if rw_cell(sheet_obj, i, 1):
            t_dict = {
                'policy': rw_cell(sheet_obj, i, 1),
                'protocol': rw_cell(sheet_obj, i, 2),
                'srcPort': rw_cell(sheet_obj, i, 3),
                'srcCidr': rw_cell(sheet_obj, i, 4),
                'destPort': rw_cell(sheet_obj, i, 5),
                'destCidr': rw_cell(sheet_obj, i, 6),
                'comment': rw_cell(sheet_obj, i, 7),
                'errors': ''
            }
            return_list.append(t_dict)
    return return_list


def print_json(output):
    print(json.dumps(output, indent=1))


def create_nw_obj_in_meraki(network, nw_objects):
    for nw_obj in nw_objects:
        response = post_network_obj(network, nw_obj)
        print(response.text)
        print(response.json())
        if response.status_code == 201:
            nw_obj['id'] = response.json()['id']
        else:
            nw_obj['id'] = response.text
    return nw_objects


def save_xls(wb_obj, file_name):
    print("saving the file to: " + file_name)
    wb_obj.save(file_name)


def add_list_to_wb(sheet_obj, list_of_dict):
    for row, line in enumerate(list_of_dict, 2):
        for col,val in enumerate(line.values(), 1):
            rw_cell(sheet_obj, row, col, write=True, value=str(val))


def create_nw_obj_groups(network, nw_object_groups):
    fresh_network_objects = get_network_objs(network)
    for nw_obj_grp in nw_object_groups:
        obj_id_list,errors = convert_obj_name_to_id(fresh_network_objects, nw_obj_grp['networkObjectNames'])
        nw_obj_grp['networkObjectIds'] = obj_id_list
        nw_obj_grp['errors'] = errors
        response = post_network_object_group(network, nw_obj_grp)
        if response.status_code == 201:
            nw_obj_grp['groupID'] = response.json()['id']
        else:
            nw_obj_grp['errors2'] = response.text
    return nw_object_groups


def create_fw_rules(network, nw_fw_rules):
    fw_rule_list = []
    for fw_rule in nw_fw_rules:
        fw_rule_list.append({
            "comment": fw_rule['comment'],
            "policy": fw_rule['policy'],
            "protocol": fw_rule['protocol'],
            "destPort": fw_rule['destPort'],
            "destCidr": fw_rule['destCidr'],
            "srcPort": fw_rule['srcPort'],
            "srcCidr": fw_rule['srcCidr'],
            "syslogEnabled": False
        })
    response = post_fw_rule(network, fw_rule_list)
    if response.status_code != 200:
        fw_rule['errors'] = response.text

    return nw_fw_rules


def convert_obj_name_to_id(network_objects, names):
    names_list = str_list_to_list(names)
    return_list = []
    errors = []
    for one_name in names_list:
        t_id = None
        for nw_obj in network_objects:
            if nw_obj['name'] == one_name:
                t_id = nw_obj['id']
        if t_id:
            return_list.append(t_id)
        else:
            errors.append(one_name)
    return (return_list, errors)


def str_list_to_list(str_list):
    t_list = str_list.split(',')
    return_list = []
    for line in t_list:
        return_list.append(line.strip())
    return return_list


def main():
    network = cli_args()

    # Set variables here if not using CLI
    # Uncomment before changing the Values
    network['api_key'] = "API Key"
    network['org_id'] = "Organization ID"
    network['net_id'] = "Network ID"
    for key,val in network.items():
        if key not in  ['output_file', 'verbose']:
            if not val:
                sys.exit('ERROR:\n\tPlease make sure all the Required arguments are used')
                continue

    xls_file = 'Meraki - Network Object Firewall Rules.xlsx'
    wb_obj = open_xls(xls_file)

    # Network Object Generation
    # Set to False if you dont want to run this
    GenerateObjects = False
    if GenerateObjects:
        print(60*"*")
        print("Starting Network Object Section")
        nw_objects = wb_read_network_objects(wb_obj['Network Object'])
        nw_objects = create_nw_obj_in_meraki(network, nw_objects)
        add_list_to_wb(wb_obj['Network Object'], nw_objects)
        print(60*"*")

    # Network Object GROUPS Generation
    # Set to False if you dont want to run this
    GenerateObjectGroups = False
    if GenerateObjectGroups:
        print(60*"*")
        print("Starting Network Object Groups Section")
        nw_object_groups = wb_read_network_object_groups(wb_obj['Network Object Groups'])
        nw_object_groups = create_nw_obj_groups(network, nw_object_groups)
        add_list_to_wb(wb_obj['Network Object Groups'], nw_object_groups)
        print(60*"*")

    # FW Rules Gnerator
    # Set to False if you dont want to run this
    GenerateFwRules = False
    if GenerateFwRules:
        print(60*"*")
        print("Starting Firewall Rules Section")
        nw_fw_rules = wb_read_fw_rules(wb_obj['Firewall Rule'])
        nw_fw_rules = create_fw_rules(network, nw_fw_rules)
        add_list_to_wb(wb_obj['Firewall Rule'], nw_fw_rules)
        print(60*"*")


    #Change output name if youd like
    output_name ="output.xlsx"
    save_xls(wb_obj, output_name)


if __name__ == '__main__':
    main()
